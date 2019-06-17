import openpyxl
try:
    from openpyxl.cell import get_column_letter,column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter,column_index_from_string
import os

os.getcwd()
wb=openpyxl.load_workbook('example.xlsx')
type(wb)

#Get sheets from the worksheet
wb.sheetnames
sheet=wb['Sheet1']
sheet.title
anotheesheet=wb.active

#Get cells from the sheets
sheet['A1']
sheet['A1'].value
c=sheet['B1']

'Row '+str(c.row)+', colum '+c.column+' is '+c.value
'Cell '+c.coordinate+' is '+c.value

#the first row and column is 1
sheet.cell(row=1,column=2)  #  'B1'
sheet.cell(row=1,column=2).value

for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)
print()
    
sheet.max_row
sheet.max_column

#Covert between column letter and number
get_column_letter(1)
column_index_from_string(get_column_letter(sheet.max_column))

#Get rows and column from the sheets
tuple(sheet['A1':'C3'])

for eachTuple in sheet['A1':'C3']: #each row is a tuple, so (A1,B1,C1),(A2,B2,C2)...
    for m in eachTuple:
        print(m.coordinate,m.value)
    print('---END OF ROW---')

for k in list(sheet.columns)[1]:   #column=list(sheet.columns)
    print(k.value)
print()

