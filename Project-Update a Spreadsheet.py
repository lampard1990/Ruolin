import openpyxl

#update price for Celery, Garlic, Lemon each for 1.19, 3.07, 1.27
wb=openpyxl.load_workbook('updatedProduceSales.xlsx')
sheet=wb['Sheet']

#store your price in a data dictionnary
price_update={'Garlic':3.07,\
             'Celery':1.19,\
             'Lemen':1.27}

for i in range(2,sheet.max_row):
    productname=sheet.cell(row=i,column=1).value
    if productname in price_update:
       sheet.cell(row=i,column=2).value=price_update[productname]


wb.save('updatedProduceSales.xlsx')
