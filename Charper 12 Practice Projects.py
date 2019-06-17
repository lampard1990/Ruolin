#Multiplication Table

import openpyxl
from openpyxl.styles import Font

wb=openpyxl.Workbook()
sheet=wb.active
bold_font = Font(bold=True)
for i in range(1,7):
    sheet.cell(row=1,column=i+1).value=i
    sheet.cell(row=1,column=i+1).font=bold_font
    sheet.cell(row=i+1,column=1).value=i
    sheet.cell(row=i+1,column=1).font=bold_font
    for j in range(1,7):
         sheet.cell(row=i+1,column=j+1).value=i*j
         

wb.save('muitiplication_table.xlsx')


#blank row insert
print('which row you want to start to insert')
N=input()
print('how many rows you want to start to insert')
M=input()
      
mysheet = openpyxl.load_workbook('BlankRowInsert.xlsx')
before_sheet = mysheet.active
before_sheet.title = 'Before'

mysheet.create_sheet(title='After')
after_sheet = mysheet['After']

n=before_sheet.max_row
for i in range(1, n+1):
      for c, cell in enumerate(before_sheet[i]): #c is for get the sequence of cells,
#start with 0, so you should +1 when paste
#for c, cell in enumerate(before_sheet[1]):
#   print(c,cell)
         if i < int(N):
             after_sheet.cell(row=i, column=c+1).value = cell.value
         else:
             after_sheet.cell(row=i+int(M), column=c+1).value = cell.value

mysheet.save('BlankRowInsert.xlsx')


#Spreadsheet cell inverter
#mysheet1 = openpyxl.load_workbook('Cell Inverter.xlsx')
#dota1 = mysheet1['Before']
#mysheet1.create_sheet(title='Inverter')
#dota2 = mysheet1['Inverter']

#k=dota1.max_row


#for i in range(1,k+1):
#      for v, mycell in enumerate(dota1[i]):
#         dota2.cell(row=v+1, column=i).value = mycell.value
        
#mysheet1.save('Cell Inverter.xlsx')




#Spreadsheet cell inverter method 2
mysheet1 = openpyxl.load_workbook('Cell Inverter.xlsx')
dota1 = mysheet1['Before']
mysheet1.create_sheet(title='Inverter')
dota2 = mysheet1['Inverter']

k=dota1.max_row
j=dota1.max_column

for i in range(1,k+1):
      for s in range(1,j+1):
         dota2.cell(row=s, column=i).value = dota1.cell(row=i, column=s).value
        
mysheet1.save('Cell Inverter.xlsx')

