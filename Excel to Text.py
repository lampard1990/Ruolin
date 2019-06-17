import openpyxl


def main():

   mysheet='spreadsheet_to_text.xlsx'

   wb = openpyxl.load_workbook(mysheet)
   sheet = wb.active

   nrows=sheet.max_row
   ncols=sheet.max_column

   for col in range(1,ncols+1):
       text_file="mysheet_"+str(col)+".txt"
       with open(text_file,'w') as file:
          for row in range(1,nrows+1):
              content=sheet.cell(row=row,column=col).value
              if content is None:
                 continue
              file.write(str(content))


if __name__ == "__main__":
   main()
