#Text to spreadsheet
import openpyxl
import os

#def main():

path=".\\12 Working With Excel Spreadsheets"

#List Comprehensions
text_files = [filename for filename in os.listdir(path) if filename.endswith(".txt")]


#text_files = []
#for filename in os.listdir(path):
#    if filename.endswith(".txt"):
#        text_files.append(filename)

#http://www.pythonforbeginners.com/basics/list-comprehensions-in-python
#new_list = []
#for i in old_list:
#    if filter(i):
#        new_list.append(expressions(i))
#You can obtain the same thing using list comprehension:
#new_list = [expression(i) for i in old_list if filter(i)]
        
wb = openpyxl.Workbook()
sheet = wb.active

#for col, text_file in enumerate(text_files):
#          with open(os.path.join(path,text_file), 'r') as filename: #filename = open("welcome.txt")
#             lines = filename.readlines() #read in one string, readline=one string each line, list object
#          print(lines)
#          for row, line in enumerate(lines):
#             sheet.cell(row=row+1, column=col+1).value = line

for i in range(len(text_files)):
         with open(os.path.join(path,text_files[i]),'r') as filename:
             content=filename.readlines()
         for j in range(len(content)):
             sheet.cell(row=j+1,column=i+1).value=content[j]

wb.save('text_to_spreadsheet.xlsx')
